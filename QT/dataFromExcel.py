import openpyxl


class GetDataFromExcel:
    """
    Excel'den istenen hastalık ve ilaçların ID numaralarını almak için oluşturulmul class
    """

    def getDiseaseFromExcel(self, diseaseType):  # Hastalıkları almak için oluşturduğumuz fonksiyon
        path = "C:\\Users\\Bilal Günden\\PycharmProjects\\Farmakogenetik\\deneme.xlsx"  # path'in çalışması için kendi local yolunuzu göstermeniz gerekiyor
        wb_obj = openpyxl.load_workbook(path)  # veri almak istediğimiz excel dosyasını ve sutünunu belirtiyoruz
        self.sheet_obj = wb_obj.active
        self.rowNumber = self.sheet_obj.max_row
        self.DiseaseList = []  # verilerimizi aktaracağımız liste
        for i in range(1, self.rowNumber):  # satır sayısı boyunca gidip beliritlen sutündaki verileri alacak
            cell_obj = self.sheet_obj.cell(row=i, column=1)
            objectValue = cell_obj.value
            if i != 1:
                modifiedValue = objectValue.split(":")
                if modifiedValue[0] == diseaseType:
                    self.DiseaseList.append(modifiedValue[1])  # listeye atama
        return self.DiseaseList

    def getDrugNameFromExcel(self):  # ilaçları almak için oluşturdğumuz fonksiyon
        path = "C:\\Users\\Bilal Günden\\PycharmProjects\\Farmakogenetik\\deneme.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        self.sheet_obj = wb_obj.active
        self.rowNumber = self.sheet_obj.max_row
        self.drugList = []
        for i in range(1, self.rowNumber):
            cell_obj = self.sheet_obj.cell(row=i, column=2)
            objectValue = cell_obj.value
            if i != 1:
                self.drugList.append(objectValue)
                print(objectValue)
        return self.drugList

